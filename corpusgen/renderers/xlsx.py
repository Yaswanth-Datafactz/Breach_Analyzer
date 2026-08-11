"""XLSX renderer (openpyxl). Table blocks only — one worksheet per Table,
row 1 = headers. Everything is written as a string so SSNs/phones keep
their leading zeros and match the manifest value byte-for-byte.
Locations: {"sheet": name, "cell": "B4"}.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from corpusgen.renderers import FIXED_DOC_DT, DocumentSpec, Table, planting_dict


def render(spec: DocumentSpec, path: Path) -> list[dict]:
    tables = [b for b in spec.blocks if isinstance(b, Table)]
    assert tables and len(tables) == len(spec.blocks), "xlsx specs are table-only"

    wb = Workbook()
    plantings: list[dict] = []
    for i, table in enumerate(tables):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = table.name
        ws.append(table.headers)
        for row in table.rows:
            ws.append([str(v) for v in row])
        for row_idx, col_idx, plant in table.cell_plants:
            assert plant.value in table.rows[row_idx][col_idx]
            cell = f"{get_column_letter(col_idx + 1)}{row_idx + 2}"
            plantings.append(planting_dict(plant, {"sheet": table.name, "cell": cell}))

    wb.properties.created = FIXED_DOC_DT
    wb.properties.modified = FIXED_DOC_DT
    wb.save(str(path))
    return plantings
