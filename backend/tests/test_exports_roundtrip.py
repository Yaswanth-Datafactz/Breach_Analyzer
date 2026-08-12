"""Export round-trips: GET /exports/exposure.csv parses back to the §2
column groups with the fixture persons' values; .xlsx re-opens via openpyxl
with identical content (both render from one iterator, so this asserts the
shared path once per format)."""

from __future__ import annotations

import csv
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.config import get_settings
from app.db.session import SessionLocal
from app.main import app
from app.services.er.persist import run_er_stage
from app.services.export import EXPORT_COLUMNS
from app.services.exposure import compute_exposure
from tests.er_scenario import build_scenario, teardown_scenario

HEADERS = {"X-API-Key": get_settings().api_key}


@pytest.fixture(scope="module")
def api():
    db = SessionLocal()
    scenario = build_scenario(db)
    run_er_stage(db, scenario.run_id)
    compute_exposure(db, scenario.run_id)
    db.commit()
    client = TestClient(app)
    try:
        yield client, scenario
    finally:
        teardown_scenario(db, scenario)
        db.close()


def test_csv_round_trip(api):
    client, scenario = api
    response = client.get(
        f"/api/v1/exports/exposure.csv?run_id={scenario.run_id}", headers=HEADERS
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "attachment" in response.headers["content-disposition"]

    rows = list(csv.reader(io.StringIO(response.text)))
    header, data = rows[0], rows[1:]
    assert header == list(EXPORT_COLUMNS)
    assert len(data) == 5  # one row per visible person

    # TWO persons print as "Robert Fournier" (the SharedName pair, kept
    # apart by design) -- the real one is the 3-mention cluster.
    robert = next(
        row
        for row in data
        if row[header.index("best_name")] == "Robert Fournier"
        and row[header.index("mention_count")] == "3"
    )

    def col(name: str) -> str:
        return robert[header.index(name)]

    # identity group
    assert "Bob Fournier [nickname]" in col("aliases")
    assert col("dob") == "1987-03-14"
    # flags group (csv serializes booleans as True/False strings)
    assert col("flag_ssn") == "True"
    assert col("flag_credit_card") == "True"
    assert col("flag_credentials") == "True"
    assert col("flag_passport") == "False"
    # evidence group
    assert col("document_count") == "3"
    assert col("mention_count") == "3"
    assert int(col("evidence_count")) == 7  # 4 ssn + 1 dob + 1 card + 1 credential
    # quality group
    assert float(col("confidence_ssn")) == pytest.approx(0.95)
    assert col("confidence_passport") == ""
    assert col("review_status") in ("auto", "needs_review", "human_confirmed")


def test_xlsx_round_trip(api):
    client, scenario = api
    response = client.get(
        f"/api/v1/exports/exposure.xlsx?run_id={scenario.run_id}", headers=HEADERS
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    sheet = workbook["exposure"]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    header, data = rows[0], rows[1:]
    assert header == list(EXPORT_COLUMNS)
    assert len(data) == 5

    robert = next(
        row
        for row in data
        if row[header.index("best_name")] == "Robert Fournier"
        and row[header.index("mention_count")] == 3
    )
    assert robert[header.index("flag_ssn")] is True  # native booleans in xlsx
    assert robert[header.index("flag_passport")] is False
    assert robert[header.index("confidence_ssn")] == pytest.approx(0.95)
    assert robert[header.index("dob")] == "1987-03-14"


def test_exports_require_auth(api):
    client, scenario = api
    assert client.get(f"/api/v1/exports/exposure.csv?run_id={scenario.run_id}").status_code == 401
    assert client.get(f"/api/v1/exports/exposure.xlsx?run_id={scenario.run_id}").status_code == 401
