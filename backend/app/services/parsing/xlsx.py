"""XLSX parsing via openpyxl, plus the deterministic header-lexicon hook.

Passages are `sheet_range` chunks of data rows ({"sheet", "row_start",
"row_end"} in real spreadsheet row numbers -- row 1 is the header row,
corpusgen's convention, so the first chunk starts at row_start=2). Every
chunk's text repeats the header line first: a chunk is the unit tier 1
sees, and column meaning must travel with it.

The header map is the phase-B2 hook for plan §9's spreadsheet rule ("the
80-person sheet costs ~ one tier-1 call, not 80"): each sheet's headers are
mapped to pii element types by an exact normalized lexicon lookup --
deterministic, no fuzz, no model -- and the map is emitted INTO the passage
locator, so the deterministic sheet extractor (detector='deterministic_sheet',
plan §4) can read column semantics straight off the passage it is anchoring
evidence to. Unmapped headers get element_type None: B2 sends only those
columns to the LLM.
"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.parsing.types import ParsedPassage, ParseResult

_CELL_SEPARATOR = " | "

# Data rows per passage: 40 keeps the bulk 80-person sheet at 2 chunks while
# staying well inside plan §9's ~2-3K-token passage-aligned chunk size.
ROWS_PER_PASSAGE = 40

# Exact-match lexicon over normalized header text -> pii_elements.element_type
# (plan §4's enum). Deliberately exact, not fuzzy: a deterministic mapping
# either knows a header or it abstains -- a near-miss goes to the LLM, never
# to a guess. Vocabulary covers corpusgen's templates (customer export:
# Name/SSN/DOB/Email/Phone/Account; cred dump: username/password/ssn) plus
# the obvious synonyms of each.
_HEADER_LEXICON: dict[str, str] = {
    "name": "name",
    "full name": "name",
    "customer name": "name",
    "employee name": "name",
    "member name": "name",
    "ssn": "ssn",
    "social security": "ssn",
    "social security number": "ssn",
    "member ssn": "ssn",
    "ssn last 4": "ssn_last4",
    "last 4": "ssn_last4",
    "dob": "dob",
    "date of birth": "dob",
    "birth date": "dob",
    "birthdate": "dob",
    "email": "email",
    "email address": "email",
    "e mail": "email",
    "phone": "phone",
    "phone number": "phone",
    "telephone": "phone",
    "mobile": "phone",
    "account": "financial_account",
    "account number": "financial_account",
    "acct": "financial_account",
    "bank account": "financial_account",
    "card": "credit_card",
    "card number": "credit_card",
    "credit card": "credit_card",
    "address": "address",
    "street address": "address",
    "home address": "address",
    "mailing address": "address",
    "drivers license": "drivers_license",
    "drivers license number": "drivers_license",
    "license number": "drivers_license",
    "dl number": "drivers_license",
    "passport": "passport",
    "passport number": "passport",
    "username": "credential",
    "user name": "credential",
    "login": "credential",
    "password": "credential",
    "diagnosis": "medical",
    "condition": "medical",
    "medical record": "medical",
    "mrn": "medical",
}


def normalize_header(header: str) -> str:
    """Lowercase, collapse everything non-alphanumeric to single spaces --
    'SSN', 'ssn ' and ' Social_Security-Number ' all normalize to lexicon
    keys ('ssn', 'social security number')."""
    return re.sub(r"[^a-z0-9]+", " ", header.lower()).strip()


def map_header(header: str) -> str | None:
    """The deterministic header-lexicon hook: element_type or None (abstain)."""
    return _HEADER_LEXICON.get(normalize_header(header))


def build_header_map(headers: list[str]) -> list[dict]:
    """[{col, letter, header, element_type}] -- col is 1-based; letter is
    the spreadsheet column letter so a B2 evidence row can cite 'B7'."""
    return [
        {
            "col": idx,
            "letter": get_column_letter(idx),
            "header": header,
            "element_type": map_header(header) if header else None,
        }
        for idx, header in enumerate(headers, start=1)
    ]


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def parse_xlsx(content: bytes) -> ParseResult:
    """Raises on corrupt input (bad zip / bad workbook) -- mapped to a
    `corrupt` quarantine by the pipeline's exception boundary."""
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        passages: list[ParsedPassage] = []
        for worksheet in workbook.worksheets:
            rows = [[_cell_str(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
            # Drop fully-empty trailing rows (read_only mode faithfully
            # reports padded dimensions on some writers).
            while rows and not any(rows[-1]):
                rows.pop()
            if not rows:
                continue

            headers = rows[0]
            header_map = build_header_map(headers)
            header_line = _CELL_SEPARATOR.join(headers)
            data_rows = rows[1:]

            if not data_rows:
                passages.append(
                    ParsedPassage(
                        kind="sheet_range",
                        locator={
                            "sheet": worksheet.title,
                            "row_start": 1,
                            "row_end": 1,
                            "header_map": header_map,
                        },
                        text=header_line,
                    )
                )
                continue

            for chunk_start in range(0, len(data_rows), ROWS_PER_PASSAGE):
                chunk = data_rows[chunk_start : chunk_start + ROWS_PER_PASSAGE]
                row_start = chunk_start + 2  # data begins at spreadsheet row 2
                row_end = row_start + len(chunk) - 1
                lines = [header_line] + [_CELL_SEPARATOR.join(row) for row in chunk]
                passages.append(
                    ParsedPassage(
                        kind="sheet_range",
                        locator={
                            "sheet": worksheet.title,
                            "row_start": row_start,
                            "row_end": row_end,
                            "header_map": header_map,
                        },
                        text="\n".join(lines),
                    )
                )
        return ParseResult(passages=passages)
    finally:
        workbook.close()
